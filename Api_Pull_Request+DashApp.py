import dash
from dash import dcc, html
from dash.dependencies import Input, Output
from zipfile import BadZipFile
import pandas as pd
import requests
import urllib3
import logging
import time
import datetime
from openpyxl import load_workbook
import os
import json
import win32com.client

# Suprimir la advertencia InsecureRequestWarning
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Configuración de logging
logging.basicConfig(
    filename='logfile.log',  # Nombre del archivo de registro
    level=logging.DEBUG,  # Nivel de registro
    format='%(asctime)s - %(levelname)s - %(message)s'  # Formato de los mensajes de registro
)

# Mapeo de serialNumber a alias
serial_to_alias = {
    'H11601054': 'RS31',
    'H11601215': 'RS47',
    'H11601217': 'RS48',
    'H11601240': 'RS49',
    'H11601241': 'RS50',
    'H11601269': 'RS51',
    'H11601270': 'RS52',
    'H11601280': 'RS53'
}

# Mapeo de alias a columnas en el archivo Excel
alias_column_map = {
    'FL 99': 'B', 'FL 102': 'C', 'FL 103': 'D', 'FL 104': 'E',
    'FL 105': 'F', 'FL 106': 'G', 'FL 107': 'H', 'FL 108': 'I',
    'FL 500': 'J', 'FL 501': 'K', 'FL160': 'L'
}

# Mapeo de alias a imágenes
imagen_por_alias = {
    'FL16': 'FL16T SIN FONDO',
    'FL': 'Forklift',
    'RS': 'Reachstaker sin fondo',
    'RTG': 'RTG SIN FONDO',
    'GR': 'MHC SIN FONDO'
}

last_api_call = 0  # Variable para almacenar el timestamp de la última llamada exitosa a la API
api_wait_time = 0  # Variable para almacenar el tiempo de espera sugerido por la API
last_update_time = "No disponible"  # Inicializar la variable global

def safe_json_loads(json_str):
    try:
        return json.loads(json_str)
    except json.JSONDecodeError as e:
        logging.error(f"Error al decodificar JSON: {e}")
        return {}
    except Exception as e:
        logging.error(f"Error desconocido: {e}")
        return {}
    
# Función personalizada para ordenar los alias
def ordenar_alias(alias):
    if isinstance(alias, str):  # Verificar que el alias sea una cadena de texto
        if alias.startswith('GR'):
            return (1, alias)
        elif alias.startswith('RTG'):
            return (2, alias)
        elif alias.startswith('RS'):
            return (3, alias)
        elif alias.startswith('FL'):
            if alias == 'FL 99':
                return (4, 'FL 00')  # Poner 'FL 99' en primer lugar dentro de los 'FL'
            else:
                return (5, alias)
    return (6, alias)  # Para cualquier otro prefijo no considerar 

def actualizar_ultima_data_valida(ntr_data, data2):
    global last_update_time
    """
    Guarda los DataFrames ntr_data y data2 en un archivo CSV.
    """
    archivo_ntr_csv = 'Last_ntr_data.csv'
    archivo_data2_csv = 'Last_data2.csv'
    try:
        # Convertir las listas a DataFrames
        df_ntr = pd.DataFrame(ntr_data)
        df2 = pd.DataFrame(data2['dataList'])
        df2['Alias'] = df2['serialNumber'].map(serial_to_alias)
        df2['Horometro'] = df2['totalRunningHours'].astype(float).round()
        df2['Fecha'] = pd.Timestamp.now().strftime("%d-%m-%Y")
        df2 = df2[['Alias', 'Horometro', 'Fecha']]

        # Serializar las columnas JSON de ntr_data
        df_ntr['gps'] = df_ntr['gps'].apply(json.dumps)
        df_ntr['entradas'] = df_ntr['entradas'].apply(json.dumps)
        df_ntr['totalizadores'] = df_ntr['totalizadores'].apply(json.dumps)

        # Añadir la fecha y hora de la última actualización
        df_ntr['last_update_time'] = datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S")

        # Guardar los DataFrames en archivos CSV separados
        df_ntr.to_csv(archivo_ntr_csv, index=False)
        df2.to_csv(archivo_data2_csv, index=False)

        logging.info(f"Datos guardados en {archivo_ntr_csv} y {archivo_data2_csv}")
        
    except Exception as e:
        logging.error("Error al guardar los archivos CSV", exc_info=True)
        raise e


def cargar_ultima_data_valida():
    """
    Carga los DataFrames desde archivos CSV.
    """
    archivo_ntr_csv = 'Last_ntr_data.csv'
    archivo_data2_csv = 'Last_data2.csv'
    if os.path.exists(archivo_ntr_csv) and os.path.exists(archivo_data2_csv):
        try:
            # Leer los DataFrames desde los archivos CSV
            df_ntr = pd.read_csv(archivo_ntr_csv)
            df2 = pd.read_csv(archivo_data2_csv)

            logging.info(f"Datos cargados desde {archivo_ntr_csv} y {archivo_data2_csv}, se carga última data")

            # Reemplazar comillas simples por comillas dobles y manejar NaN en df_ntr
            df_ntr['gps'] = df_ntr['gps'].fillna('{}').str.replace("'", '"')
            df_ntr['entradas'] = df_ntr['entradas'].fillna('{}').str.replace("'", '"')
            df_ntr['totalizadores'] = df_ntr['totalizadores'].fillna('{}').str.replace("'", '"')

            # Deserializar columnas JSON en df_ntr
            df_ntr['gps'] = df_ntr['gps'].apply(safe_json_loads)
            df_ntr['entradas'] = df_ntr['entradas'].apply(safe_json_loads)
            df_ntr['totalizadores'] = df_ntr['totalizadores'].apply(safe_json_loads)

            # Convertir el DataFrame en una lista de diccionarios
            ntr_data = df_ntr.to_dict(orient='records')

            # Convertir df2 en una lista de diccionarios
            data2 = df2.to_dict(orient='records')

            # Extraer la fecha y hora de la última actualización
            if 'last_update_time' in df_ntr.columns:
                global last_update_time
                last_update_time = df_ntr['last_update_time'].iloc[0]

            return ntr_data, data2
        except Exception as e:
            logging.error("Error al cargar los archivos CSV", exc_info=True)
            raise e
    else:
        logging.warning(f"Archivos {archivo_ntr_csv} o {archivo_data2_csv} no encontrados.")
        return [], []  # Retorna dos listas vacías si los archivos no existen


# Carga inicial de los datos
last_valid_data = cargar_ultima_data_valida()

def obtener_datos():
    global last_api_call, api_wait_time

    url = "https://api.caesistemas.com.ar/v1/ntr.php"
    params = {"api_key": "xxx"}

    try:
        current_time = time.time()
        if current_time < last_api_call + api_wait_time:
            logging.warning("Llamadas a la API demasiado frecuentes. Esperando...")
            return None
            
        logging.info("Enviando solicitud a la API...")
        response = requests.get(url, params=params, verify=False, timeout=10)

        if response.status_code == 200:
            logging.info("Solicitud exitosa")
            data = response.json()
            estado = data.get("estado", {})
            codigo = estado.get("codigo")

            if codigo == -2:
                logging.warning("Esperando...")
                mensaje = estado.get("mensaje")
                tiempo_espera = int(mensaje.split(" ")[1])
                logging.info(f"Esperando {tiempo_espera} segundos antes de intentar nuevamente...")
                last_api_call = time.time()
                api_wait_time = tiempo_espera
                return None
            elif codigo == 0:
                logging.info("Datos disponibles para procesar")
                last_api_call = time.time()
                api_wait_time = 0  # Resetear el tiempo de espera si la llamada fue exitosa
                return data["datos"]["ntr"]
            else:
                mensaje = estado.get("mensaje")
                raise Exception(f"Error en la API: {mensaje}")
        else:
            raise Exception("Error al obtener los datos de la API:", response.status_code)
        
    except requests.exceptions.RequestException as e:
        logging.error(f"Error de conexión: {e}")        
    except Exception as e:
        logging.error("Error inesperado en obtener_datos", exc_info=True)
        return None
    
def obtener_datos2():
    url2 = "https://cloud-api.digi.kalmarglobal.com/runningHours"
    headers = {"X-API-KEY": "xxx"}

    try:
        response2 = requests.get(url2, headers=headers, verify=False)

        if response2.status_code == 200:
            logging.info("Solicitud exitosa a la segunda API")
            return response2.json()
        else:
            raise Exception("Error al obtener los datos de la API de running hours:", response2.status_code)
    except requests.exceptions.RequestException as e:
        logging.error(f"Error de conexión a la segunda API: {e}")
        return None
    
def procesar_datos2(data2):
    try:
        if 'dataList' in data2:  # Verifica si data2 tiene la clave 'dataList'
            df2 = pd.DataFrame(data2['dataList'])
        else:
            df2 = pd.DataFrame(data2)

        df2['Alias'] = df2['serialNumber'].map(serial_to_alias)
        df2['Horometro'] = df2['totalRunningHours'].astype(float).round()
        df2['Fecha'] = pd.Timestamp.now().strftime("%d-%m-%Y")
        df2 = df2[['Alias', 'Horometro', 'Fecha']]
        return df2
    except Exception as e:
        logging.error("Error en procesar_datos2", exc_info=True)
        return pd.DataFrame()

def obtener_imagen(alias):
    for prefijo, imagen in imagen_por_alias.items():
        if alias.startswith(prefijo):
            return imagen
    return 'Imagen no encontrada'  # Si no se encuentra ningún prefijo

def actualizar_excel_forklift(df_filtrado):
    try:
        archivo_excel = 'Forklift_Horometros.xlsx'
        ruta_completa = os.path.abspath(archivo_excel)
        logging.info(f'Intentando cargar el archivo Excel: {ruta_completa}')

        wb = load_workbook(archivo_excel)
        ws = wb.active

        fecha_actual = datetime.datetime.now().strftime('%d/%m/%Y')

        for _, row in df_filtrado.iterrows():
            alias = row['Alias']
            horometro = row['Horometro']

            horometro = round(horometro)

            if alias in alias_column_map:
                columna = alias_column_map[alias]
                ws[f'{columna}9'] = horometro
                ws[f'{columna}10'] = fecha_actual

        wb.save(archivo_excel)
        logging.info(F'Datos actualizados en {ruta_completa}.')

    except FileNotFoundError:
        logging.error(f'Archivo no encontrado: {ruta_completa}')
    except BadZipFile:
        logging.error(f'El archivo no es un archivo zip válido o está corrupto: {ruta_completa}')
    except Exception as e:
        logging.error("Error inesperado en actualizar_excel_forklift", exc_info=True)
        raise e  

def procesar_datos():
    global last_valid_data, last_update_time, api_wait_time, last_api_call
    try:
        ntr_data = obtener_datos()
        data2 = obtener_datos2()

        df2 = procesar_datos2(data2)

        if ntr_data is None:
            if last_valid_data is not None:
                ntr_data, data2 = last_valid_data
                if df2.empty:
                    df2 = procesar_datos2(last_valid_data[1])
                logging.info(f"Usando la última data válida.{last_valid_data}")
            else:
                raise Exception("No hay datos válidos disponibles.")
        else:
            logging.info("Actualizando la última data válida.")
            last_update_time = datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S")
            last_api_call = time.time()  # Actualizar el tiempo de la última llamada API
            last_valid_data = (ntr_data, data2)
            actualizar_ultima_data_valida(ntr_data, data2)
            # Actualizar el tiempo de última actualización

    except Exception as e:
        logging.error("Error en procesar_datos", exc_info=True)
        raise e

    alias_list = [item['alias'] for item in ntr_data]
    horometro_list = [item['totalizadores']['horometro'] for item in ntr_data]

    df = pd.DataFrame({'Alias': alias_list, 'Horometro': horometro_list})

    # Unir los datos de las dos APIs
    df = pd.concat([df, df2]).drop_duplicates(subset='Alias', keep='last')

    ajustes = {
        'FL 103': 4142,
        'FL 104': 60,
        'FL 105': 27,
        'FL 106': 6323,
        'FL 107': 33,
        'FL 99': 4752,
        'FL 501': 40,
        'GR08': 2994,
        'GR09': 2240,
        'GR10': -33,
        'RTG02': 20598,
        'RTG11': 154,
        'RTG16': 547,
        'RTG17': 39,
        'RTG20': 47928,
        'RTG24': 31763,
        'RTG26': 13517,
        'RTG31': 20138,
        'RS 39': 38894,
        'RS 40': 201,
    }

    df['Horometro'] = df.apply(lambda row: row['Horometro'] + ajustes.get(row['Alias'], 0), axis=1)
    df['Fecha'] = pd.Timestamp.now().strftime("%d-%m-%Y")

    # Aplicar la función obtener_imagen usando lambda
    df['Imagen'] = df['Alias'].apply(lambda x: obtener_imagen(x))

    # Convertir todos los valores en la columna Alias a cadenas de texto
    df['Alias'] = df['Alias'].astype(str)

    df_filtrado = df[df['Alias'].str.startswith(('RS', 'RTG', 'GR', 'FL'))]
    # Aplicar la función de orden personalizada
    df_filtrado = df_filtrado.sort_values(by='Alias', key=lambda x: x.map(ordenar_alias))

    df_horometros = pd.read_csv('horometros_anteriores.csv')
    logging.info("Archivo CSV cargado exitosamente para comparar valores semanales.")

    for _, row in df_filtrado.iterrows():
        alias = row['Alias']
        horometro_actual = row['Horometro']
        if alias.startswith('FL') or alias.startswith('RS') or alias.startswith('GR') or alias.startswith('RTG'):
            if alias in df_horometros['alias'].values:
                horometro_anterior = df_horometros.loc[df_horometros['alias'] == alias, 'horometro'].values[0]
                timestamp_anterior = df_horometros.loc[df_horometros['alias'] == alias, 'timestamp'].values[0]
                tiempo_transcurrido = datetime.datetime.now().timestamp() - timestamp_anterior
                logging.info(f"Tiempo transcurrido: {tiempo_transcurrido}")
                if tiempo_transcurrido >= 603000: #604800
                    diferencia = horometro_actual - horometro_anterior
                    logging.info(f"Alias: {alias}, Horometro semana anterior: {horometro_anterior}, Horometro semana actual: {horometro_actual}, Diferencia: {diferencia}")
                    df_horometros.loc[df_horometros['alias'] == alias, 'horometro'] = horometro_actual
                    df_horometros.loc[df_horometros['alias'] == alias, 'timestamp'] = datetime.datetime.now().timestamp()
                    df_horometros.loc[df_horometros['alias'] == alias, 'diferencia'] = diferencia
            else:
                new_row = pd.DataFrame({
                    'alias': [alias],
                    'horometro': [horometro_actual],
                    'timestamp': [datetime.datetime.now().timestamp()],
                    'diferencia': [0]
                })
                df_horometros = pd.concat([df_horometros, new_row], ignore_index=True)

    # Guardar el DataFrame actualizado como un archivo Excel
    df_horometros.to_excel('horometros_anteriores.xlsx', index=True)
    # Y CSV
    df_horometros.to_csv('horometros_anteriores.csv', index=False)
    df_filtrado.to_csv('df_filtrado.csv', index=False)
    df_filtrado.to_excel('df_filtrado.xlsx', index=False)
    logging.info("DataFrame filtrado guardado como 'df_filtrado.csv'.")

    actualizar_excel_forklift(df_filtrado)
    logging.info("Proceso completado.")

    return df_filtrado

# Crear la aplicación Dash
app = dash.Dash(__name__)

# Layout de la aplicación
app.layout = html.Div(style={'backgroundColor': '#C0C0C0'},children=[
    html.H1(className='header', children=[
        html.H1(children='Horómetros de Equipos', style={'text-align': 'center','background-color': '#C0C0C0'}),
    ]),

    dcc.Interval(
        id='interval-component',
        interval= 3600*1000,  # Actualizar cada 1 hora
        n_intervals=0
    ),

    # Añadir dcc.Store para almacenar los datos
    dcc.Store(id='stored-data'),

    html.Div(id='live-update-text', style={'text-align': 'center', 'font-size': '24px', 'font-weight': 'bold','background-color': '#C0C0C0'}),
    html.Div(id='images', style={'display': 'flex', 'flex-wrap': 'wrap', 'justify-content': 'center','background-color': '#C0C0C0'})
])

@app.callback(
    [Output('live-update-text', 'children'),
     Output('images', 'children'),
     Output('stored-data', 'data')],
    [Input('interval-component', 'n_intervals')]
)

def display_data(n):
    global last_update_time, api_wait_time, last_api_call
    try:
        df_filtrado = procesar_datos()

        last_update_text = f'Última actualización: {last_update_time}'

        # Calcular el tiempo restante para la próxima actualización
        current_time = time.time()
        tiempo_restante = max(0, last_api_call + api_wait_time - current_time)
        minutos_restantes = tiempo_restante // 60 if tiempo_restante > 0 else api_wait_time // 60
        update_info_text = f'{last_update_text}   -   Próxima actualización aproximada en: {int(minutos_restantes)} minutos'

        # Crear elementos de imagen y texto para cada equipo
        images = []
        categories = {'GR': 'GRÚAS', 'RTG': 'RTG', 'RS': 'REACHSTAKER', 'FL': 'FORKLIFT'}
        added_categories = set()
        
        for _, equipo in df_filtrado.iterrows():

            alias = equipo['Alias']
            category_prefix = next((prefix for prefix in categories if alias.startswith(prefix)), None)

            if category_prefix and category_prefix not in added_categories:
                images.append(html.H2(children=categories[category_prefix], style={'width': '100%', 'text-align': 'center','font-size': '32px','font-weight': 'bold','margin-top': '20px','background-color': '#C0C0C0'}))
                added_categories.add(category_prefix)

            images.append(html.Div(className='equipment-container', children=[
                html.Img(src=f'/assets/{equipo["Imagen"]}.png', className='equipment-image', style={'width': '180px', 'height': '100px'}),
                html.Div(f'{equipo["Alias"]} - Horómetro: {equipo["Horometro"]}', className='equipment-text', style={'font-size': '20px', 'font-weight': 'bold'})
            ], style={'margin': '20px', 'padding': '10px', 'border': '1px solid #ddd', 'border-radius': '10px', 'text-align': 'center','background-color': '#ADD8E6'}))

        # Aplicar estilos CSS al texto de última actualización
        last_update_element = html.Div(update_info_text, style={'font-size': '30px', 'font-weight': 'bold'})

        # Devolver el texto, las imágenes y los datos almacenados
        return last_update_element, images, df_filtrado.to_dict('records')

    except Exception as e:
        logging.error("Error al procesar datos", exc_info=True)
        return html.P("Error al procesar datos."), [], []

# Ejecutar la aplicación Dash
if __name__ == '__main__':
    app.run_server(debug=True)
